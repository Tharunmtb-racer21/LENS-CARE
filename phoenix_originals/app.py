from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, send_file
from datetime import datetime, timedelta
import os
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import colors
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import openpyxl
from openpyxl import Workbook
import psycopg2
import psycopg2.extras
from contextlib import contextmanager

app = Flask(__name__)
app.secret_key = 'lens_care_secret_key_2026_change_in_production'
app.permanent_session_lifetime = timedelta(days=30)

# Excel files
USERS_FILE = 'users.xlsx'
PRESCRIPTIONS_FILE = 'prescriptions.xlsx'
EYE_TESTS_FILE = 'eye_tests.xlsx'

# ─────────────────────────────────────────────
#  SQLite — mirrors every Excel write
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────
#  POSTGRESQL  (Supabase / Render Postgres)
#  Reads DATABASE_URL from environment — set this in Render dashboard.
#  Falls back gracefully so local Excel-only mode still works.
# ─────────────────────────────────────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL', '')

# Render injects "postgres://" but psycopg2 needs "postgresql://"
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

@contextmanager
def get_db():
    """Open a PostgreSQL connection; yields cursor; auto-commit or rollback."""
    if not DATABASE_URL:
        # No database configured — yield None so callers can skip gracefully
        yield None
        return
    conn = psycopg2.connect(DATABASE_URL)
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        yield cur
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def init_db():
    """Create tables and seed doctor account (safe to call every startup)."""
    if not DATABASE_URL:
        print("⚠️  DATABASE_URL not set — skipping PostgreSQL init (Excel-only mode)")
        return
    with get_db() as db:
        # Users table
        db.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id              SERIAL      PRIMARY KEY,
                username        TEXT        NOT NULL UNIQUE,
                email           TEXT        NOT NULL UNIQUE,
                password_hash   TEXT        NOT NULL,
                full_name       TEXT,
                date_of_birth   TEXT,
                phone           TEXT,
                role            TEXT        NOT NULL DEFAULT 'patient',
                created_at      TEXT        NOT NULL,
                onboarding_done INTEGER     NOT NULL DEFAULT 0
            )
        """)
        # Prescriptions table
        db.execute("""
            CREATE TABLE IF NOT EXISTS prescriptions (
                id                SERIAL  PRIMARY KEY,
                user_id           INTEGER NOT NULL,
                username          TEXT,
                full_name         TEXT,
                prescription_date TEXT,
                age               TEXT,
                od_sph  REAL, od_cyl  REAL, od_axis REAL, od_add  REAL,
                os_sph  REAL, os_cyl  REAL, os_axis REAL, os_add  REAL,
                diagnosis         TEXT,
                lens_type         TEXT,
                lens_power        TEXT,
                lens_material     TEXT,
                focal_length      TEXT,
                od_se             REAL,
                os_se             REAL,
                colorblind_score  TEXT,
                colorblind_type   TEXT,
                acuity_score      TEXT,
                notes             TEXT
            )
        """)
        # Eye tests table
        db.execute("""
            CREATE TABLE IF NOT EXISTS eye_tests (
                id              SERIAL  PRIMARY KEY,
                user_id         INTEGER NOT NULL,
                username        TEXT,
                test_date       TEXT,
                test_type       TEXT,
                score           TEXT,
                result          TEXT,
                colorblind_type TEXT
            )
        """)
        # Seed doctor account
        db.execute("SELECT id FROM users WHERE username=%s", ('doctor',))
        if not db.fetchone():
            db.execute("""
                INSERT INTO users (username,email,password_hash,full_name,
                                   date_of_birth,phone,role,created_at,onboarding_done)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, ('doctor', 'doctor@lenscare.com',
                  generate_password_hash('doctor123'),
                  'Dr. Admin', '1980-01-01', '1234567890',
                  'doctor', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 1))
            print("✅ PostgreSQL: doctor account seeded")


def init_excel_files():
    """Initialize Excel files if they don't exist"""
    # Users file
    if not os.path.exists(USERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(['id', 'username', 'email', 'password_hash', 'full_name', 'date_of_birth', 'phone', 'role', 'created_at', 'onboarding_done'])
        
        # Add pre-saved doctor account
        doctor_password = generate_password_hash('doctor123')
        ws.append([1, 'doctor', 'doctor@lenscare.com', doctor_password, 'Dr. Admin', '1980-01-01', '1234567890', 'doctor', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), True])
        
        wb.save(USERS_FILE)
        print("✅ Users file created with doctor account")
        print("   Username: doctor | Password: doctor123")
    
    # Prescriptions file
    if not os.path.exists(PRESCRIPTIONS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Prescriptions"
        ws.append(['id', 'user_id', 'username', 'full_name', 'prescription_date', 'age', 
                   'od_sph', 'od_cyl', 'od_axis', 'od_add', 'os_sph', 'os_cyl', 'os_axis', 'os_add',
                   'diagnosis', 'lens_type', 'lens_power', 'lens_material', 'focal_length',
                   'od_se', 'os_se', 'colorblind_score', 'colorblind_type', 'acuity_score', 'notes'])
        wb.save(PRESCRIPTIONS_FILE)
        print("✅ Prescriptions file created")
    
    # Eye tests file
    if not os.path.exists(EYE_TESTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "EyeTests"
        ws.append(['id', 'user_id', 'username', 'test_date', 'test_type', 'score', 'result', 'colorblind_type'])
        wb.save(EYE_TESTS_FILE)
        print("✅ Eye tests file created")

init_excel_files()

def migrate_users_file():
    """Add onboarding_done column to existing users.xlsx if missing"""
    try:
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        if 'onboarding_done' not in header:
            ws.cell(row=1, column=10, value='onboarding_done')
            for row in ws.iter_rows(min_row=2):
                role = row[7].value if len(row) > 7 else 'patient'
                row[9].value = True if role == 'doctor' else False
            wb.save(USERS_FILE)
            print("✅ Migrated users.xlsx — added onboarding_done column")
        wb.close()
    except Exception as e:
        print(f"Migration skipped: {e}")

migrate_users_file()
init_db()   # ← initialise SQLite (safe to call every startup)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def calculate_lens_specifications(data):
    """Calculate lens specifications"""
    try:
        age = int(data.get('age', 30))
        od_sph = float(data.get('od_sph', 0))
        od_cyl = float(data.get('od_cyl', 0))
        os_sph = float(data.get('os_sph', 0))
        os_cyl = float(data.get('os_cyl', 0))
        
        od_se = od_sph + (od_cyl / 2)
        os_se = os_sph + (os_cyl / 2)
        avg_power = (abs(od_se) + abs(os_se)) / 2
        
        focal_length_od = round(100 / od_se, 2) if od_se != 0 else "∞"
        focal_length_os = round(100 / os_se, 2) if os_se != 0 else "∞"
        
        diagnosis = []
        lens_nature = []
        lens_type = "Unknown"
        
        if od_se < -0.5 or os_se < -0.5:
            diagnosis.append("Myopia")
            lens_nature.append("Concave Lens")
        if od_se > 0.5 or os_se > 0.5:
            diagnosis.append("Hyperopia")
            lens_nature.append("Convex Lens")
        if abs(od_cyl) >= 0.5 or abs(os_cyl) >= 0.5:
            diagnosis.append("Astigmatism")
            lens_type = "Cylindrical/Toric Lens"
        if age >= 40:
            od_add = float(data.get('od_add', 0))
            os_add = float(data.get('os_add', 0))
            if od_add > 0 or os_add > 0:
                diagnosis.append("Presbyopia")
                lens_type = "Progressive/Bifocal Lens"
        
        if not diagnosis:
            diagnosis.append("Normal Vision")
            lens_type = "Plano Lens"
        elif lens_type == "Unknown":
            lens_type = " + ".join(lens_nature) if lens_nature else "Single Vision Lens"
        
        if avg_power < 2.0:
            lens_material = "CR-39 Plastic (n=1.498)"
        elif avg_power < 4.0:
            lens_material = "Polycarbonate (n=1.586)"
        elif avg_power < 6.0:
            lens_material = "High-index 1.67"
        else:
            lens_material = "High-index 1.74"
        
        recommendations = []
        if abs(od_se) > 6.0 or abs(os_se) > 6.0:
            recommendations.append("High myopia - Consider anti-reflective coating")
        if age < 18 and (od_se < -0.5 or os_se < -0.5):
            recommendations.append("Consider myopia control lenses")
        if age >= 60:
            recommendations.append("Regular eye exams recommended")
        
        return {
            'diagnosis': ', '.join(diagnosis),
            'lens_type': lens_type,
            'lens_power': f"OD: {od_se:+.2f}D, OS: {os_se:+.2f}D",
            'lens_material': lens_material,
            'focal_length': f"OD: {focal_length_od} cm, OS: {focal_length_os} cm",
            'recommendations': recommendations,
            'od_se': od_se,
            'os_se': os_se,
            'avg_power': avg_power
        }
    except Exception as e:
        return {
            'diagnosis': 'Error in calculation',
            'lens_type': 'N/A',
            'lens_power': 'N/A',
            'lens_material': 'N/A',
            'focal_length': 'N/A',
            'recommendations': [],
            'od_se': 0,
            'os_se': 0,
            'avg_power': 0
        }

@app.route('/')
def index():
    # Always show the public landing page
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            data = request.json
            username = data.get('username')
            email = data.get('email')
            password = data.get('password')
            full_name = data.get('full_name')
            date_of_birth = data.get('date_of_birth')
            phone = data.get('phone', '')
            
            # Check if username exists
            wb = openpyxl.load_workbook(USERS_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == username or row[2] == email:
                    wb.close()
                    return jsonify({'success': False, 'message': 'Username or email already exists'})
            
            # Add new user
            user_id = ws.max_row
            password_hash = generate_password_hash(password)
            ws.append([user_id, username, email, password_hash, full_name, date_of_birth, phone, 'patient', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), False])
            wb.save(USERS_FILE)
            wb.close()

            # ── PostgreSQL mirror ──
            try:
                with get_db() as db:
                    if db is not None:
                        db.execute("""
                            INSERT INTO users
                                (username,email,password_hash,full_name,date_of_birth,
                                 phone,role,created_at,onboarding_done)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            ON CONFLICT (username) DO NOTHING
                        """, (username, email, password_hash, full_name, date_of_birth,
                              phone, 'patient', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 0))
            except Exception as sql_err:
                print(f"PostgreSQL register warning: {sql_err}")
            
            # Auto-login the new user so they land on /onboarding directly
            session.permanent = True
            session['user_id']  = user_id
            session['username'] = username
            session['full_name'] = full_name
            session['role']     = 'patient'
            return jsonify({'success': True, 'message': 'Registration successful', 'redirect': '/onboarding'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            data = request.json
            username = data.get('username')
            password = data.get('password')
            
            wb = openpyxl.load_workbook(USERS_FILE)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == username:
                    if check_password_hash(row[3], password):
                        session.permanent = True
                        session['user_id'] = row[0]
                        session['username'] = row[1]
                        session['full_name'] = row[4]
                        session['role'] = row[7]
                        wb.close()
                        role = row[7]
                        # row[9] = onboarding_done flag (True/False/None)
                        onboarding_done = row[9] if len(row) > 9 else False
                        if role == 'doctor':
                            redirect_url = '/doctor_dashboard'
                        elif onboarding_done:
                            redirect_url = '/dashboard'
                        else:
                            redirect_url = '/onboarding'
                        return jsonify({'success': True, 'message': 'Login successful', 'redirect': redirect_url, 'role': role})
            
            wb.close()
            return jsonify({'success': False, 'message': 'Invalid credentials'})
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))  # Always go back to public landing page

@app.route('/dashboard')
@login_required
def dashboard():
    wb_prescriptions = openpyxl.load_workbook(PRESCRIPTIONS_FILE)
    ws_prescriptions = wb_prescriptions.active
    
    wb_tests = openpyxl.load_workbook(EYE_TESTS_FILE)
    ws_tests = wb_tests.active
    
    wb_users = openpyxl.load_workbook(USERS_FILE)
    ws_users = wb_users.active
    
    # Get user info
    user = None
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {
                'id': row[0], 'username': row[1], 'email': row[2],
                'full_name': row[4], 'date_of_birth': row[5], 'phone': row[6], 'role': row[7]
            }
            break
    
    # Get prescriptions (all if doctor, own if patient)
    prescriptions = []
    for row in ws_prescriptions.iter_rows(min_row=2, values_only=True):
        if session.get('role') == 'doctor' or row[1] == session['user_id']:
            prescriptions.append({
                'id': row[0], 'user_id': row[1], 'username': row[2], 'full_name': row[3],
                'prescription_date': row[4], 'age': row[5],
                'od_sph': row[6], 'od_cyl': row[7], 'od_axis': row[8], 'od_add': row[9],
                'os_sph': row[10], 'os_cyl': row[11], 'os_axis': row[12], 'os_add': row[13],
                'diagnosis': row[14], 'lens_type': row[15], 'lens_power': row[16],
                'lens_material': row[17], 'focal_length': row[18],
                'od_se': row[19], 'os_se': row[20]
            })
    
    prescriptions = sorted(prescriptions, key=lambda x: x['prescription_date'], reverse=True)
    
    # Get eye tests
    eye_tests = []
    for row in ws_tests.iter_rows(min_row=2, values_only=True):
        if session.get('role') == 'doctor' or row[1] == session['user_id']:
            eye_tests.append({
                'id': row[0], 'user_id': row[1], 'username': row[2],
                'test_date': row[3], 'test_type': row[4], 'score': row[5], 'result': row[6]
            })
    
    eye_tests = sorted(eye_tests, key=lambda x: x['test_date'], reverse=True)[:5]
    
    wb_users.close()
    wb_prescriptions.close()
    wb_tests.close()
    
    return render_template('dashboard.html', user=user, prescriptions=prescriptions, eye_tests=eye_tests)

@app.route('/eye_test')
@login_required
def eye_test():
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    user = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {
                'id': row[0], 'username': row[1],
                'full_name': row[4], 'role': row[7]
            }
            break
    wb.close()
    return render_template('eye_test.html', user=user)

@app.route('/save_eye_test', methods=['POST'])
@login_required
def save_eye_test():
    try:
        data = request.json
        wb = openpyxl.load_workbook(EYE_TESTS_FILE)
        ws = wb.active
        
        test_id = ws.max_row
        
        if data.get('colorblind_score') is not None:
            ws.append([
                test_id, session['user_id'], session['username'],
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'color_blindness', data['colorblind_score'], 
                data.get('colorblind_result', ''),
                data.get('colorblind_type', 'Normal')
            ])
        
        if data.get('acuity_score'):
            ws.append([
                test_id + 1, session['user_id'], session['username'],
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'visual_acuity', data['acuity_score'], 
                data.get('acuity_result', ''), ''
            ])
        
        wb.save(EYE_TESTS_FILE)
        wb.close()

        # ── PostgreSQL mirror ──
        try:
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            with get_db() as db:
                if db is not None:
                    if data.get('colorblind_score') is not None:
                        db.execute("""
                            INSERT INTO eye_tests
                                (user_id,username,test_date,test_type,score,result,colorblind_type)
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                        """, (session['user_id'], session['username'], now,
                              'color_blindness', data['colorblind_score'],
                              data.get('colorblind_result',''),
                              data.get('colorblind_type','Normal')))
                    if data.get('acuity_score'):
                        db.execute("""
                            INSERT INTO eye_tests
                                (user_id,username,test_date,test_type,score,result,colorblind_type)
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                        """, (session['user_id'], session['username'], now,
                              'visual_acuity', data['acuity_score'],
                              data.get('acuity_result',''), ''))
        except Exception as sql_err:
            print(f"PostgreSQL eye_test warning: {sql_err}")

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/prescription')
@login_required
def prescription():
    return render_template('prescription.html')

@app.route('/submit_prescription', methods=['POST'])
@login_required
def submit_prescription():
    try:
        data = request.json
        specs = calculate_lens_specifications(data)
        
        wb = openpyxl.load_workbook(PRESCRIPTIONS_FILE)
        ws = wb.active
        
        prescription_id = ws.max_row
        
        ws.append([
            prescription_id, session['user_id'], session['username'], session['full_name'],
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'), data.get('age'),
            data.get('od_sph'), data.get('od_cyl'), data.get('od_axis'), data.get('od_add', ''),
            data.get('os_sph'), data.get('os_cyl'), data.get('os_axis'), data.get('os_add', ''),
            specs['diagnosis'], specs['lens_type'], specs['lens_power'], specs['lens_material'],
            specs['focal_length'], specs['od_se'], specs['os_se'],
            data.get('colorblind_score', ''), data.get('colorblind_type', ''),
            data.get('acuity_score', ''), data.get('notes', '')
        ])
        
        wb.save(PRESCRIPTIONS_FILE)
        wb.close()

        # ── PostgreSQL mirror ──
        try:
            with get_db() as db:
                if db is not None:
                    db.execute("""
                        INSERT INTO prescriptions
                            (user_id,username,full_name,prescription_date,age,
                             od_sph,od_cyl,od_axis,od_add,
                             os_sph,os_cyl,os_axis,os_add,
                             diagnosis,lens_type,lens_power,lens_material,focal_length,
                             od_se,os_se,colorblind_score,colorblind_type,acuity_score,notes)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (session['user_id'], session['username'], session['full_name'],
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S'), data.get('age'),
                          data.get('od_sph'), data.get('od_cyl'), data.get('od_axis'), data.get('od_add',''),
                          data.get('os_sph'), data.get('os_cyl'), data.get('os_axis'), data.get('os_add',''),
                          specs['diagnosis'], specs['lens_type'], specs['lens_power'],
                          specs['lens_material'], specs['focal_length'],
                          specs['od_se'], specs['os_se'],
                          data.get('colorblind_score',''), data.get('colorblind_type',''),
                          data.get('acuity_score',''), data.get('notes','')))
        except Exception as sql_err:
            print(f"PostgreSQL prescription warning: {sql_err}")

        session['current_prescription_id'] = prescription_id
        session['analysis'] = specs
        session['patient_data'] = data
        
        return jsonify({'success': True, 'message': 'Prescription saved successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/analysis')
@login_required
def analysis():
    specs = session.get('analysis', {})
    patient_data = session.get('patient_data', {})
    
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    user = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {'id': row[0], 'username': row[1], 'full_name': row[4]}
            break
    wb.close()
    
    return render_template('analysis.html', specs=specs, patient_data=patient_data, user=user)

@app.route('/prescription_history/<int:prescription_id>')
@login_required
def prescription_history(prescription_id):
    wb = openpyxl.load_workbook(PRESCRIPTIONS_FILE)
    ws = wb.active
    
    prescription = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == prescription_id:
            if session.get('role') == 'doctor' or row[1] == session['user_id']:
                prescription = {
                    'id': row[0], 'user_id': row[1], 'username': row[2], 'full_name': row[3],
                    'prescription_date': row[4], 'age': row[5],
                    'od_sph': row[6], 'od_cyl': row[7], 'od_axis': row[8], 'od_add': row[9],
                    'os_sph': row[10], 'os_cyl': row[11], 'os_axis': row[12], 'os_add': row[13],
                    'diagnosis': row[14], 'lens_type': row[15], 'lens_power': row[16],
                    'lens_material': row[17], 'focal_length': row[18],
                    'od_se': row[19], 'os_se': row[20],
                    'colorblind_score': row[21], 'colorblind_type': row[22],
                    'acuity_score': row[23], 'notes': row[24]
                }
                break
    
    wb.close()
    
    if not prescription:
        flash('Prescription not found', 'danger')
        return redirect(url_for('dashboard'))
    
    return render_template('prescription_detail.html', prescription=prescription)

@app.route('/view_all_records')
@login_required
def view_all_records():
    wb_prescriptions = openpyxl.load_workbook(PRESCRIPTIONS_FILE)
    ws_prescriptions = wb_prescriptions.active
    
    wb_tests = openpyxl.load_workbook(EYE_TESTS_FILE)
    ws_tests = wb_tests.active
    
    # Get all prescriptions (doctor sees all, patient sees own)
    prescriptions = []
    for row in ws_prescriptions.iter_rows(min_row=2, values_only=True):
        if session.get('role') == 'doctor' or row[1] == session['user_id']:
            prescriptions.append({
                'id': row[0], 'user_id': row[1], 'username': row[2], 'full_name': row[3],
                'prescription_date': row[4], 'age': row[5],
                'od_sph': row[6], 'od_cyl': row[7], 'od_axis': row[8], 'od_add': row[9],
                'os_sph': row[10], 'os_cyl': row[11], 'os_axis': row[12], 'os_add': row[13],
                'diagnosis': row[14], 'lens_type': row[15], 'lens_power': row[16],
                'lens_material': row[17], 'focal_length': row[18],
                'od_se': row[19], 'os_se': row[20]
            })
    
    prescriptions = sorted(prescriptions, key=lambda x: x['prescription_date'], reverse=True)
    
    # Get all eye tests
    eye_tests = []
    for row in ws_tests.iter_rows(min_row=2, values_only=True):
        if session.get('role') == 'doctor' or row[1] == session['user_id']:
            eye_tests.append({
                'id': row[0], 'user_id': row[1], 'username': row[2],
                'test_date': row[3], 'test_type': row[4], 'score': row[5], 'result': row[6]
            })
    
    eye_tests = sorted(eye_tests, key=lambda x: x['test_date'], reverse=True)
    
    wb_prescriptions.close()
    wb_tests.close()
    
    return render_template('view_all_records.html', prescriptions=prescriptions, eye_tests=eye_tests)


@app.route('/download_report/<int:prescription_id>')
@login_required
def download_report(prescription_id):
    try:
        wb = openpyxl.load_workbook(PRESCRIPTIONS_FILE)
        ws = wb.active
        
        prescription = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == prescription_id:
                prescription = row
                break
        
        wb.close()
        
        if not prescription:
            return "Prescription not found", 404
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        title = Paragraph(f"<b>LENS CARE - Prescription Report</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        info_data = [
            ['Patient Name:', prescription[3]],
            ['Prescription Date:', str(prescription[4])]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 11),
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 11),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        rx_data = [
            ['Eye', 'Sphere', 'Cylinder', 'Axis', 'Add'],
            ['OD', str(prescription[6]), str(prescription[7]), str(prescription[8]), str(prescription[9] or '-')],
            ['OS', str(prescription[10]), str(prescription[11]), str(prescription[12]), str(prescription[13] or '-')]
        ]
        
        rx_table = Table(rx_data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch])
        rx_table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 11),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(rx_table)
        
        doc.build(story)
        buffer.seek(0)
        
        return send_file(buffer, as_attachment=True, 
                        download_name=f"prescription_{prescription_id}.pdf", 
                        mimetype='application/pdf')
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/api/prescription_trend/<int:user_id>')
@login_required
def prescription_trend(user_id):
    # Allow int or string comparison (session may store as either)
    sid = session.get('user_id')
    try:
        sid = int(sid)
    except (TypeError, ValueError):
        pass

    if sid != user_id and session.get('role') != 'doctor':
        return jsonify({'error': 'Unauthorized'}), 403

    try:
        wb = openpyxl.load_workbook(PRESCRIPTIONS_FILE)
        ws = wb.active

        prescriptions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row[1] is user_id stored in Excel — compare as int
            row_uid = row[1]
            try:
                row_uid = int(row_uid)
            except (TypeError, ValueError):
                pass

            if row_uid == user_id and len(row) >= 21:
                prescriptions.append({
                    'prescription_date': str(row[4]) if row[4] else '',
                    'od_se':   row[19], 'os_se':   row[20],
                    'od_sph':  row[6],  'os_sph':  row[10],
                    'od_cyl':  row[7],  'os_cyl':  row[11],
                    'od_axis': row[8],  'os_axis': row[12],
                    'od_add':  row[9],  'os_add':  row[13],
                    'lens_power': row[16],
                    'diagnosis': row[14] or ''
                })

        wb.close()

        # Sort safely — skip rows where date is empty/None
        prescriptions = sorted(
            [p for p in prescriptions if p['prescription_date']],
            key=lambda x: x['prescription_date']
        )

        def safe_float(v):
            try:
                return float(v) if v is not None and v != '' else 0.0
            except (TypeError, ValueError):
                return 0.0

        return jsonify({
            'dates':     [p['prescription_date'] for p in prescriptions],
            'od_se':     [safe_float(p['od_se'])   for p in prescriptions],
            'os_se':     [safe_float(p['os_se'])   for p in prescriptions],
            'od_sph':    [safe_float(p['od_sph'])  for p in prescriptions],
            'os_sph':    [safe_float(p['os_sph'])  for p in prescriptions],
            'od_cyl':    [safe_float(p['od_cyl'])  for p in prescriptions],
            'os_cyl':    [safe_float(p['os_cyl'])  for p in prescriptions],
            'od_axis':   [safe_float(p['od_axis']) for p in prescriptions],
            'os_axis':   [safe_float(p['os_axis']) for p in prescriptions],
            'od_add':    [safe_float(p['od_add'])  for p in prescriptions],
            'os_add':    [safe_float(p['os_add'])  for p in prescriptions],
            'diagnoses': [p['diagnosis']           for p in prescriptions]
        })

    except Exception as e:
        return jsonify({'error': str(e), 'dates': [], 'od_se': [], 'os_se': [],
                        'od_sph': [], 'os_sph': [], 'od_cyl': [], 'os_cyl': [],
                        'od_axis': [], 'os_axis': [], 'od_add': [], 'os_add': [],
                        'diagnoses': []}), 200



@app.route('/eye_insights')
@login_required
def eye_insights():
    """Eye analytics dashboard with 6 charts"""
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    user = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {
                'id': row[0], 'username': row[1],
                'full_name': row[4], 'role': row[7]
            }
            break
    wb.close()
    return render_template('eye_insights.html', user=user)

@app.route('/onboarding')
@login_required
def onboarding():
    """Post-login onboarding questionnaire — skipped if already completed"""
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    user = None
    onboarding_done = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {'id': row[0], 'username': row[1], 'full_name': row[4], 'role': row[7]}
            onboarding_done = row[9] if len(row) > 9 else False
            break
    wb.close()

    # Already done — go straight to dashboard
    if onboarding_done:
        return redirect(url_for('dashboard'))

    return render_template('onboarding.html', user=user)

@app.route('/save_onboarding', methods=['POST'])
@login_required
def save_onboarding():
    """Mark onboarding complete in Excel so it never shows again for this user"""
    try:
        data = request.json
        session['onboarding'] = data
        session['onboarding_done'] = True

        # Persist flag to Excel — column J (index 10, 1-based) = onboarding_done
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == session['user_id']:
                # Ensure column J exists (auto-extends if sheet is narrow)
                row[9].value = True
                break
        wb.save(USERS_FILE)
        wb.close()

        # ── PostgreSQL mirror ──
        try:
            with get_db() as db:
                if db is not None:
                    db.execute(
                        "UPDATE users SET onboarding_done=1 WHERE username=%s",
                        (session['username'],)
                    )
        except Exception as sql_err:
            print(f"PostgreSQL onboarding warning: {sql_err}")

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/lens_frames')
@login_required
def lens_frames():
    """Lens frames and recommendations page"""
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    user = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {'id': row[0], 'username': row[1], 'full_name': row[4], 'role': row[7]}
            break
    wb.close()
    return render_template('lens_frames.html', user=user)


@app.route('/face_analyser')
@login_required
def face_analyser():
    """Face shape analyser — uses MediaPipe in-browser, no server processing needed"""
    with get_db() as db:
        pass  # get_db used for potential future save; no DB read needed here
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    user = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {
                'id': row[0], 'username': row[1],
                'full_name': row[4], 'role': row[7]
            }
            break
    wb.close()
    return render_template('face_analyser.html', user=user)


@app.route('/frame_creator')
@login_required
def frame_creator():
    """Frame Creator — virtual try-on studio with live MediaPipe tracking"""
    wb = openpyxl.load_workbook(USERS_FILE)
    ws = wb.active
    user = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {'id':row[0],'username':row[1],'full_name':row[4],'role':row[7]}
            break
    wb.close()
    return render_template('frame_creator.html', user=user)

@app.route('/doctor_dashboard')
@login_required
def doctor_dashboard_route():
    """Doctor-only dashboard"""
    if session.get('role') != 'doctor':
        return redirect(url_for('dashboard'))
    
    wb_prescriptions = openpyxl.load_workbook(PRESCRIPTIONS_FILE)
    ws_prescriptions = wb_prescriptions.active
    wb_users = openpyxl.load_workbook(USERS_FILE)
    ws_users = wb_users.active

    user = None
    for row in ws_users.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user_id']:
            user = {'id': row[0], 'username': row[1], 'full_name': row[4], 'role': row[7]}
            break

    all_prescriptions = []
    for row in ws_prescriptions.iter_rows(min_row=2, values_only=True):
        all_prescriptions.append({
            'id': row[0], 'user_id': row[1], 'username': row[2], 'full_name': row[3],
            'prescription_date': row[4], 'diagnosis': row[14]
        })

    all_prescriptions = sorted(all_prescriptions, key=lambda x: x['prescription_date'] or '', reverse=True)
    
    # Unique patients
    seen = set()
    patients = []
    for p in all_prescriptions:
        if p['user_id'] not in seen:
            seen.add(p['user_id'])
            patients.append(p)

    wb_prescriptions.close()
    wb_users.close()

    return render_template('doctor_dashboard.html',
        user=user,
        session=session,
        all_prescriptions=all_prescriptions,
        patients=patients,
        total_patients=len(patients),
        total_visits=len(all_prescriptions)
    )


if __name__ == '__main__':
    print("=" * 70)
    print("👁️  LENS CARE - Interactive Optical Lens Advisor")
    print("=" * 70)
    print(f"\n✅ Excel files initialized")
    print(f"📁 Data persists in Excel files:")
    print(f"   - {USERS_FILE}")
    print(f"   - {PRESCRIPTIONS_FILE}")
    print(f"   - {EYE_TESTS_FILE}")
    print(f"   - PostgreSQL via DATABASE_URL env var")
    print(f"\n🩺 PRE-SAVED DOCTOR LOGIN:")
    print(f"   Username: doctor")
    print(f"   Password: doctor123")
    print(f"   (Can view ALL patients)")
    print("\n" + "=" * 70)
    print("🌐 SERVER STARTING...")
    print("=" * 70)
    print("\n🎯 Open your browser:")
    print("   👉 http://localhost:5000")
    print("\n📝 Features:")
    print("   ✅ Excel storage (data persists)")
    print("   ✅ Doctor login pre-saved")
    print("   ✅ Complete eye tests")
    print("   ✅ 4 dashboard charts")
    print("   ✅ Sunset gradient theme")
    print("\n💡 Press Ctrl+C to stop")
    print("=" * 70 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)
